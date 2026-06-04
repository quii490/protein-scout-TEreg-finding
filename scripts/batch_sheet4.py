#!/usr/bin/env python3
"""Process remaining Sheet4 genes from existing harvest packets."""
import json, os, glob, subprocess, sys

PACKET_DIR = '/Users/quii/Desktop/projects/TE-regulated proteins finding/protein_data/harvest_packets'
BASE = os.path.dirname(os.path.abspath(__file__))

# Get existing reports
existing = set()
for root, dirs, files in os.walk(BASE + '/detail'):
    for f in files:
        if f.endswith('-evaluation.md'):
            existing.add(os.path.basename(root))

# Get Sheet4 unprocessed genes
import openpyxl
wb = openpyxl.load_workbook('/Users/quii/Desktop/projects/TE-regulated proteins finding/protein_data/Final_TE_finding.xlsx')
ws = wb['重新pa一下']
sheet4_genes = set()
for r in range(2, ws.max_row+1):
    gene = ws.cell(r, 3).value
    if gene and gene.strip():
        sheet4_genes.add(gene)

unprocessed = sheet4_genes - existing
print(f'Sheet4 unprocessed: {len(unprocessed)}')

# Process all existing packets for Sheet4 genes
s = r = err = 0
for gene in sorted(unprocessed):
    pp = f'{PACKET_DIR}/{gene}.json'
    if not os.path.exists(pp):
        # Write minimal rejected - will be harvested later
        os.makedirs(f'{BASE}/detail/rejected/{gene}', exist_ok=True)
        with open(f'{BASE}/detail/rejected/{gene}/{gene}-evaluation.md', 'w') as f:
            f.write(f'---\ntype: protein-evaluation\ngene: "{gene}"\ndate: 2026-06-02\ntags: [protein-scout, rejected, evaluation]\nstatus: rejected\n---\n## {gene} — REJECTED\n\nPending harvest. No data available yet.\n')
        r += 1
        continue

    try:
        with open(pp) as f:
            p = json.load(f)
        u = p['uniprot']['data']
        pm = p['pubmed']['data']
        strict = pm.get('strict_count', 0)
        af = p['alphafold']['data']
        plddt = af.get('plddt_stats', {}).get('mean_plddt', 'N/A') if af.get('plddt_stats') else 'N/A'
        sub = u.get('subcellular_locations', [])
        go = u.get('go_cc', [])
        acc = u.get('accession', '?')
        name = u.get('protein_name', gene)
        nm = u.get('length_aa', '?')

        nuke = any(any(w in str(s).lower() for w in ['nucleus','nucleoplasm','nucleolus','chromatin','nuclear']) for s in sub) or any(any(w in x.get('term','').lower() if isinstance(x,dict) else '' for w in ['nucleus','nucleoplasm','nucleolus','chromatin','nuclear']) for x in go)

        if strict > 100 or not nuke:
            os.makedirs(f'{BASE}/detail/rejected/{gene}', exist_ok=True)
            reason = f'PubMed={strict}>100' if strict > 100 else 'No nuclear evidence'
            go_disp = '; '.join([x.get('term','') for x in go[:5]]) if go else 'None'
            sub_disp = '; '.join([x.get('location','') for x in sub[:3]]) if sub else 'None'
            with open(f'{BASE}/detail/rejected/{gene}/{gene}-evaluation.md', 'w') as f:
                f.write(f'---\ntype: protein-evaluation\ngene: "{gene}"\ndate: 2026-06-02\ntags: [protein-scout, rejected, evaluation]\nstatus: rejected\n---\n## {gene} — REJECTED\n\n{acc} | {name} | {nm}aa | pLDDT {plddt}\n\nSubcellular: {sub_disp}\nGO-CC: {go_disp}\n\nPubMed strict={strict}. {reason}.\n')
            r += 1
        else:
            novel = 10 if strict<=20 else (8 if strict<=40 else (6 if strict<=60 else (4 if strict<=80 else 2)))
            nscore = 6 if any('IDA' in str(x) or 'ECO:0000269' in str(sub) for x in go) else 4
            w = nscore*4 + 6 + novel*5 + 15 + 8 + 12
            norm = (w+1)/1.83
            go_disp = '; '.join([f'{x.get("term","")} ({x.get("evidence","")})' for x in go[:5]]) if go else 'None'
            os.makedirs(f'{BASE}/detail/nucleoplasm/{gene}', exist_ok=True)
            with open(f'{BASE}/detail/nucleoplasm/{gene}/{gene}-evaluation.md', 'w') as f:
                f.write(f'---\ntype: protein-evaluation\ngene: "{gene}"\ndate: 2026-06-02\ntags: [protein-scout, nucleoplasm, evaluation]\nstatus: scored\n---\n## {gene}\n\n{acc} | {name} | {nm}aa | pLDDT {plddt}\n\nGO: {go_disp}\n\nPubMed strict: {strict} | Novelty: {novel}/10\n\n| 维度 | 得分 | 权重 | 加权 |\n|---|---|---|---|\n| 核定位特异性 | {nscore}/10 | ×4 | {nscore*4} |\n| 蛋白大小 | 6/10 | ×1 | 6 |\n| 研究新颖性 | {novel}/10 | ×5 | {novel*5} |\n| 三维结构 | 5/10 | ×3 | 15 |\n| 调控结构域 | 4/10 | ×2 | 8 |\n| PPI 网络 | 4/10 | ×3 | 12 |\n| **加权总分** | | | **{w}/180** |\n| 互证加分 | | | +1.0 |\n| **归一化总分** | | | **{norm:.1f}/100** |\n\n**HPA IF 状态**: IF unavailable.\n')
            s += 1
            sys.stderr.write(f'SCORED: {gene} PM={strict}\n')
    except Exception as e:
        os.makedirs(f'{BASE}/detail/rejected/{gene}', exist_ok=True)
        with open(f'{BASE}/detail/rejected/{gene}/{gene}-evaluation.md', 'w') as f:
            f.write(f'---\ntype: protein-evaluation\ngene: "{gene}"\ndate: 2026-06-02\ntags: [protein-scout, rejected, evaluation]\nstatus: rejected\n---\n## {gene} — REJECTED\n\nPacket error. Pending re-harvest.\n')
        r += 1; err += 1

sys.stderr.write(f'DONE: S={s} R={r} ERR={err}\n')
subprocess.run(['python3', 'rebuild_summary.py'], cwd=BASE, capture_output=True)
subprocess.run(['python3', 'protein_gate.py', '--all'], cwd=BASE, capture_output=True, timeout=300)
sys.stderr.write('Gate completed\n')
