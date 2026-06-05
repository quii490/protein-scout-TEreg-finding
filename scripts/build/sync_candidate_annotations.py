#!/usr/bin/env python3
from __future__ import annotations

import argparse
import csv
import json
import re
from collections import defaultdict
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[2]
INDEX = ROOT / "docs" / "data" / "protein_report_index.json"
ANNOTATIONS = ROOT / "data" / "manual" / "web_annotations.tsv"
ENGINEERING = ROOT / "audit_work" / "engineering"
DEFAULT_CANDIDATE_DIR = Path("/Users/quii/Desktop/projects/TE-regulated proteins finding/candidates")

FIELDS = ["gene", "report_path", "row_class", "note", "priority"]
MISSING_FIELDS = ["gene", "sources"]
HEADER_HINTS = {
    "gene", "gene name", "gene names (primary)", "first_gene_name",
    "candidates", "less studied", "analyze current data",
    "bio-id", "bioid", "symbol",
}
BAD_TOKENS = {
    "GENE", "GENES", "GENENAME", "ENTRY", "CANDIDATES", "PROTEIN", "PROTEINS",
    "OE", "KI", "BIO", "ID", "NA", "NONE", "CONTROL", "GUIDE", "RNA", "SEQ",
    "DNA", "DOMAIN", "COMPLEX", "BLANK", "NC", "HDAC", "THO", "TSN", "BAH",
    "AGO", "PHB", "FAMB", "SUPERKILLER", "VASCULIN",
}


def load_web_genes() -> set[str]:
    data = json.loads(INDEX.read_text(encoding="utf-8"))
    return {rec["gene"].upper() for rec in data.get("records", []) if rec.get("gene")}


def normalize_token(value: object) -> list[str]:
    text = str(value or "").strip()
    if not text:
        return []
    # Do not mine gene symbols from long descriptions. This prevents false
    # positives such as MYB from "MYBBP1A binds Myb proto-oncogene protein".
    if len(text) > 80 or len(re.findall(r"\w+", text)) > 6:
        return []
    text = re.sub(r"[（(].*?[）)]", "", text)
    text = text.replace("\u3000", " ")
    out: list[str] = []
    for raw in re.split(r"[,;，；\s\n\t]+", text):
        raw = raw.strip().strip("'\"`")
        if not raw:
            continue
        expanded = [raw]
        slash = re.fullmatch(r"([A-Za-z]+)([0-9A-Za-z]+)(?:/([0-9A-Za-z]+))+", raw)
        if slash:
            prefix = slash.group(1)
            parts = raw[len(prefix):].split("/")
            expanded = [prefix + part for part in parts]
        elif "/" in raw:
            expanded = [part for part in raw.split("/") if part]
        for item in expanded:
            item = re.sub(r"[^A-Za-z0-9_.-]", "", item).upper()
            if re.fullmatch(r"[ACGTU]{10,}", item):
                continue
            if not re.fullmatch(r"[A-Z][A-Z0-9_.-]{1,30}", item):
                continue
            if item in BAD_TOKENS:
                continue
            out.append(item)
    return out


def header_like(value: object) -> bool:
    text = str(value or "").strip().lower()
    return text in HEADER_HINTS or bool(re.fullmatch(r".*(?:gene|candidate|bio-?id|symbol).*", text))


def candidate_columns(ws) -> tuple[int, list[int]]:
    best_row = 1
    cols: list[int] = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=min(ws.max_row, 12), values_only=True), start=1):
        found = [idx for idx, value in enumerate(row) if header_like(value)]
        if found:
            best_row = row_idx
            cols = found
            break
    if not cols:
        # No reliable header: assume candidate symbols are in the first few
        # identifier-like columns, not in long free-text annotation columns.
        return 0, list(range(min(ws.max_column, 3)))
    return best_row, cols


def extract_xlsx(path: Path) -> dict[str, set[str]]:
    hits: dict[str, set[str]] = defaultdict(set)
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    for ws in wb.worksheets:
        header_row, cols = candidate_columns(ws)
        for row in ws.iter_rows(min_row=header_row + 1, max_row=ws.max_row, values_only=True):
            values = [row[idx] for idx in cols if idx < len(row)]
            for value in values:
                for gene in normalize_token(value):
                    hits[gene].add(f"{path.name}:{ws.title}")
    return hits


def load_existing(valid_genes: set[str]) -> dict[tuple[str, str], dict[str, str]]:
    if not ANNOTATIONS.exists():
        return {}
    rows: dict[tuple[str, str], dict[str, str]] = {}
    with ANNOTATIONS.open("r", encoding="utf-8", newline="") as f:
        reader = csv.DictReader(f, delimiter="\t")
        for row in reader:
            gene = (row.get("gene") or "").strip()
            report_path = (row.get("report_path") or "").strip()
            priority = (row.get("priority") or "").strip().lower()
            row_class = (row.get("row_class") or "").strip().lower()
            if gene and gene.upper() in valid_genes and (priority != "candidate" or row_class == "danger"):
                rows[(gene.upper(), report_path)] = {k: row.get(k, "") for k in FIELDS}
    return rows


def write_annotations(hits: dict[str, set[str]], web_genes: set[str]) -> dict[str, set[str]]:
    ANNOTATIONS.parent.mkdir(parents=True, exist_ok=True)
    rows = load_existing(web_genes)
    web_hits = {gene: sources for gene, sources in hits.items() if gene in web_genes}
    for gene, sources in web_hits.items():
        key = (gene, "")
        row = rows.get(key, {field: "" for field in FIELDS})
        row["gene"] = gene
        row["report_path"] = row.get("report_path", "")
        if row.get("row_class") != "danger":
            row["row_class"] = "warning"
        if not row.get("priority") or row.get("priority") == "candidate" or row.get("note", "").startswith("候选表命中"):
            row["note"] = ""
        row["priority"] = row.get("priority") or "candidate"
        rows[key] = row
    ordered = sorted(rows.values(), key=lambda r: (r.get("gene", "").upper(), r.get("report_path", "")))
    with ANNOTATIONS.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=FIELDS, delimiter="\t")
        writer.writeheader()
        writer.writerows(ordered)
    return web_hits


def write_missing_candidates(hits: dict[str, set[str]], web_genes: set[str]) -> Path:
    missing = {gene: sources for gene, sources in hits.items() if gene not in web_genes}
    path = ENGINEERING / "candidate_genes_missing_from_web_index.tsv"
    with path.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=MISSING_FIELDS, delimiter="\t")
        writer.writeheader()
        for gene in sorted(missing):
            writer.writerow({"gene": gene, "sources": "; ".join(sorted(missing[gene]))})
    return path


def main() -> None:
    parser = argparse.ArgumentParser(description="Highlight proteins found in candidate spreadsheets.")
    parser.add_argument("--candidate-dir", type=Path, default=DEFAULT_CANDIDATE_DIR)
    args = parser.parse_args()
    web_genes = load_web_genes()
    hits: dict[str, set[str]] = defaultdict(set)
    for path in sorted(args.candidate_dir.glob("*.xlsx")):
        for gene, sources in extract_xlsx(path).items():
            hits[gene].update(sources)
    web_hits = write_annotations(hits, web_genes)
    missing_path = write_missing_candidates(hits, web_genes)
    ENGINEERING.mkdir(parents=True, exist_ok=True)
    report = ENGINEERING / "candidate_annotation_sync_summary.md"
    report.write_text(
        "\n".join([
            "# Candidate Annotation Sync Summary",
            "",
            f"- Candidate dir: {args.candidate_dir}",
            f"- Extracted candidate genes: {len(hits)}",
            f"- Highlighted genes in web index: {len(web_hits)}",
            f"- Missing candidate genes from web index: {len(hits) - len(web_hits)}",
            f"- Output: {ANNOTATIONS.relative_to(ROOT)}",
            f"- Missing candidates TSV: {missing_path.relative_to(ROOT)}",
            "",
            "## Genes",
            "",
            "\n".join(f"- {gene}: {', '.join(sorted(sources))}" for gene, sources in sorted(web_hits.items())),
        ]),
        encoding="utf-8",
    )
    print(f"highlighted={len(web_hits)} extracted={len(hits)} missing={len(hits)-len(web_hits)} output={ANNOTATIONS} missing_tsv={missing_path} summary={report}")


if __name__ == "__main__":
    if ROOT.name != "protein-scout-TEreg-finding":
        raise SystemExit(f"Unexpected project root: {ROOT}")
    main()
